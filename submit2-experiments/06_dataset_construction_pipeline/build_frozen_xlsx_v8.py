"""
Generate the corresponding xlsx file for dataset_items_final_v2.csv.
"""
import openpyxl
import pandas as pd

SRC_XLSX = "Dataset_items_v7_final.xlsx"
OUT_XLSX = "Dataset_items_v8_final.xlsx"
NEW_CSV = "dataset_items_final_v2.csv"

new_df = pd.read_csv(NEW_CSV)
new_df.columns = [c.strip() for c in new_df.columns]
text_by_id = dict(zip(new_df["item_id"], new_df["item_text"]))

wb = openpyxl.load_workbook(SRC_XLSX)
ws = wb.active

header = [c.value for c in ws[1]]
id_col = header.index("item_id") + 1
text_col = header.index("item_text") + 1

n_changed = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    item_id = row[id_col - 1].value
    if item_id in text_by_id:
        new_text = text_by_id[item_id]
        old_text = row[text_col - 1].value
        if old_text != new_text:
            row[text_col - 1].value = new_text
            n_changed += 1

wb.save(OUT_XLSX)
print(f"OK, Saved {OUT_XLSX}, {n_changed} item_text cells updated")
