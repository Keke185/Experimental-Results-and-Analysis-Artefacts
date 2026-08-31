"""
    Final annotation records for 50 weakly aligned entries in the dataset Dataset_items_v4_final.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

SRC_PATH = "/sessions/practical-upbeat-wright/mnt/outputs/Dataset_items_v4_final.xlsx"
OUT_PATH = "/sessions/practical-upbeat-wright/mnt/outputs/Weakly_Aligned_Annotation_FINAL.xlsx"

ANNOTATIONS = {
    "Q_101": ("WA-F", 1, "Supporting (single)", "Troubleshooting",
              "Troubleshooting methodology for production issues with no obvious single point of failure; a genuine troubleshooting support capability, not scenario dressing.", "Keep - WA-F"),
    "Q_102": ("WA-F", 1, "Supporting (single)", "API integration",
              "[New item] Third-party identity-verification API integration, replacing the original entry-level CRUD item.", "Keep - WA-F (rewritten)"),
    "Q_103": ("WA-F", 1, "Supporting (single)", "Troubleshooting / performance optimisation",
              "Slow-query optimisation is a genuine performance-troubleshooting support capability; generic but technically specific and assessable.", "Keep - WA-F"),
    "Q_104": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (distributed state)",
              "Presence tracking for a meeting/communication platform; a scenario-embedded distributed state-consistency problem that does not involve signalling-protocol detail.", "Keep - WA-D"),
    "Q_105": ("WA-F", 1, "Supporting (single)", "Distributed service design (load balancing)",
              "Load-balancing principles and health checks; a supporting capability within distributed service design, not scenario dressing.", "Keep - WA-F"),
    "Q_106": ("WA-F", 1, "Supporting (single)", "Cross-team technical coordination",
              "Delivering a technical feature in collaboration with PM/QA/infra/frontend; maps clearly to the recognised \"cross-team technical coordination\" category.", "Keep - WA-F"),
    "Q_107": ("WA-F", 1, "Supporting (single)", "Troubleshooting / performance optimisation (caching)",
              "Caching strategy and invalidation risk; a performance/reliability support capability, not scenario dressing.", "Keep - WA-F"),
    "Q_108": ("WA-F", 2, "Supporting (multiple)", "Troubleshooting + observability (incident management)",
              "A complete incident-response workflow (triage/escalation/rollback/post-mortem), covering both the troubleshooting and observability support capabilities.", "Keep - WA-F"),
    "Q_109": ("WA-F", 1, "Supporting (single)", "API integration",
              "[New item] API design serving both internal and external consumers (version evolution/compatibility), replacing the original PR-review soft-skill item.", "Keep - WA-F (rewritten)"),
    "Q_110": ("WA-F", 2, "Supporting (multiple)", "Observability + troubleshooting (performance baseline)",
              "Establishing a performance baseline, load testing, bottleneck identification and reporting; covers both the observability and troubleshooting support capabilities.", "Keep - WA-F"),
    "Q_111": ("WA-F", 1, "Supporting (single)", "Security & authentication",
              "[New item] Dual human/machine authentication model design (password + API key), replacing the original requirements-clarification soft-skill item.", "Keep - WA-F (rewritten)"),
    "Q_112": ("WA-F", 1, "Supporting (single)", "Security & authentication",
              "[New item] Sensitive-data-leak incident response, replacing the original risk-communication soft-skill item.", "Keep - WA-F (rewritten)"),
    "Q_113": ("WA-D", 2, "Adjacent domain (single)", "API integration / API design",
              "Meeting-scheduling API design (idempotency, time zones, version evolution); scenario-embedded but requires only generic API-design capability, no core protocol knowledge.", "Keep - WA-D"),
    "Q_114": ("WA-D", 2, "Adjacent domain (single)", "Async task processing (webhook delivery)",
              "Reliable delivery of meeting-lifecycle webhooks; scenario-embedded but fundamentally a generic asynchronous-event-delivery capability.", "Keep - WA-D"),
    "Q_115": ("WA-D", 2, "Adjacent domain (single)", "Cloud deployment / security & authentication (storage + access)",
              "Recording-file metadata service (storage lifecycle, access control); scenario-embedded but generic storage-and-permissions design.", "Keep - WA-D"),
    "Q_116": ("WA-D", 2, "Adjacent domain (single)", "Async task processing (async workflow)",
              "Post-meeting asynchronous processing pipeline (transcription/thumbnails/compliance scanning); scenario-embedded but a generic asynchronous task-orchestration capability.", "Keep - WA-D"),
    "Q_117": ("WA-F", 1, "Supporting (single)", "Async task processing",
              "[New item] Asynchronous generation of a long-running report, replacing the original codebase-onboarding soft-skill item.", "Keep - WA-F (rewritten)"),
    "Q_118": ("WA-F", 1, "Supporting (single)", "Troubleshooting",
              "Attributing responsibility for and diagnosing a post-change production issue; the core is still troubleshooting methodology, with communication only a secondary component.", "Keep - WA-F"),
    "Q_119": ("WA-D", 2, "Adjacent domain (single)", "Async task processing / API integration (notifications)",
              "Multi-channel delivery of meeting invites/reminders; scenario-embedded but a generic notification-system-design capability.", "Keep - WA-D"),
    "Q_120": ("WA-D", 2, "Adjacent domain (single)", "API integration (third-party integration)",
              "Third-party calendar-service integration (OAuth/rate limiting/webhook sync); scenario-embedded but a generic third-party API-integration capability.", "Keep - WA-D"),
    "Q_121": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (cross-region)",
              "Cross-region session-state availability and consistency; scenario-embedded but a generic distributed-systems-design capability.", "Keep - WA-D"),
    "Q_122": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (concurrency control)",
              "Atomic enforcement of a maximum concurrent-participant count; scenario-embedded but a generic distributed concurrency-control capability.", "Keep - WA-D"),
    "Q_123": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (state management)",
              "Waiting-room state-machine design; scenario-embedded but a generic state-management/queueing-design capability.", "Keep - WA-D"),
    "Q_124": ("WA-D", 2, "Adjacent domain (single)", "Security & authentication (token service)",
              "Issuance and revocation of short-lived meeting tokens; scenario-embedded but a generic identity token-service-design capability.", "Keep - WA-D"),
    "Q_125": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design / API integration (rate limiting)",
              "Distributed rate limiting for meeting create/join APIs; scenario-embedded but a generic rate-limiting-architecture capability.", "Keep - WA-D"),
    "Q_126": ("WA-D", 1, "Adjacent domain (single, weak)", "Cloud deployment (feature rollout)",
              "Gradual rollout / feature flags; its link to the meeting scenario is only a surface wrapper and the in-domain anchor is weak, confirmed by the user to be kept.", "Keep - WA-D (confirmed)"),
    "Q_127": ("WA-D", 2, "Adjacent domain (single)", "Security & authentication / observability (audit trail)",
              "Audit logging for meeting-management operations; scenario-embedded but a generic audit-and-compliance-logging capability.", "Keep - WA-D"),
    "Q_128": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (reconnect/recovery)",
              "State recovery after disconnection/reconnection; scenario-embedded but a generic distributed state-coordination capability.", "Keep - WA-D"),
    "Q_129": ("WA-F", 1, "Supporting (single)", "Async task processing",
              "[New item] Reliability design for scheduled-job scheduling, replacing the original generic-Java-exception-handling item.", "Keep - WA-F (rewritten)"),
    "Q_130": ("WA-D", 2, "Adjacent domain (single)", "Observability (health monitoring)",
              "Health monitoring across the session/notification/recording-metadata services; scenario-embedded but a generic observability capability.", "Keep - WA-D"),
    "Q_131": ("WA-F", 1, "Supporting (single)", "Observability (audit log schema)",
              "Audit-log database schema design; a genuine observability/compliance support capability, not scenario dressing.", "Keep - WA-F"),
    "Q_132": ("WA-F", 1, "Supporting (single)", "Distributed service design (database selection)",
              "Relational vs. non-relational database selection; a genuine data-architecture support capability, not scenario dressing.", "Keep - WA-F"),
    "Q_133": ("WA-F", 1, "Supporting (single)", "Cloud deployment / distributed service design (schema migration)",
              "Safe production schema migration; a genuine deployment-safety support capability, not scenario dressing.", "Keep - WA-F"),
    "Q_134": ("WA-F", 1, "Supporting (single)", "Distributed service design (read replicas)",
              "Using read replicas to improve performance; a genuine data-replication support capability, not scenario dressing.", "Keep - WA-F"),
    "Q_135": ("WA-F", 1, "Supporting (single)", "Troubleshooting (connection pool)",
              "Diagnosing database connection-pool exhaustion; a genuine troubleshooting support capability, not scenario dressing.", "Keep - WA-F"),
    "Q_136": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design / security & authentication (multi-tenancy)",
              "Multi-tenant isolation and quotas for a meeting backend; scenario-embedded but a generic multi-tenant-architecture capability.", "Keep - WA-D"),
    "Q_137": ("WA-F", 1, "Supporting (single)", "Troubleshooting (performance regression)",
              "Systematic investigation of response-time regression; a genuine troubleshooting support capability, not scenario dressing.", "Keep - WA-F"),
    "Q_138": ("WA-D", 2, "Adjacent domain (single)", "Async task processing / observability (event stream)",
              "Event-stream design for participant state changes; scenario-embedded but a generic event-driven-architecture capability.", "Keep - WA-D"),
    "Q_139": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (coordination state)",
              "Coordinating screen-share control (explicitly excludes media-transport detail); scenario-embedded but a generic distributed-coordination capability.", "Keep - WA-D"),
    "Q_140": ("WA-F", 1, "Supporting (single)", "Cloud deployment (cloud migration)",
              "[New item] On-premises-to-cloud migration plan design, replacing the original enterprise-user-management CRUD item (removal of the original item confirmed by the user).", "Keep - WA-F (rewritten)"),
    "Q_141": ("WA-F", 1, "Supporting (single)", "Distributed service design (data synchronisation)",
              "[New item] Data synchronisation and conflict resolution between two services, replacing the original engineer-disagreement soft-skill item.", "Keep - WA-F (rewritten)"),
    "Q_142": ("WA-F", 1, "Supporting (single)", "Observability (ops handover docs)",
              "Operations handover documentation (monitoring/alerting/rollback procedure); a genuine observability/operations support capability.", "Keep - WA-F"),
    "Q_143": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (overload protection)",
              "Gateway overload protection ahead of a large meeting; scenario-embedded but a generic rate-limiting/degradation-architecture capability.", "Keep - WA-D"),
    "Q_144": ("WA-F", 1, "Supporting (single)", "Troubleshooting / cross-team coordination (incident comms)",
              "Cross-team communication during an incident; consistent in spirit with Q_106/Q_108, confirmed by the user to be kept.", "Keep - WA-F (confirmed)"),
    "Q_145": ("WA-F", 1, "Supporting (single)", "Cross-team technical coordination (cross-team dependency)",
              "[New item] Coordinating a breaking API change across teams, replacing the original cost-benefit-judgement soft-skill item.", "Keep - WA-F (rewritten)"),
    "Q_146": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (session management)",
              "Backend design for meeting-session create/join/leave; scenario-embedded but a generic session-state-management capability, no signalling-protocol detail.", "Keep - WA-D"),
    "Q_147": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (data modelling)",
              "Metadata model design for real-time audio/video sessions; scenario-embedded but a generic data-modelling capability.", "Keep - WA-D"),
    "Q_148": ("WA-D", 2, "Adjacent domain (single)", "API integration / distributed service design (protocol gateway)",
              "Dual-protocol session-gateway design; confirmed by the user not to involve concrete protocol (SIP/H.323) conversion detail, kept as WA-D.", "Keep - WA-D (confirmed)"),
    "Q_149": ("WA-D", 2, "Adjacent domain (single)", "Observability (RTC monitoring)",
              "Monitoring-metric design for a real-time session backend; scenario-embedded but a generic observability capability.", "Keep - WA-D"),
    "Q_150": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design / cloud deployment (scaling)",
              "Scaling design for a communication platform from hundreds to tens of thousands of concurrent users; scenario-embedded but a generic horizontal-scaling capability.", "Keep - WA-D"),
}

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WAD_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
WAF_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
REWRITE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def main():
    src_wb = openpyxl.load_workbook(SRC_PATH)
    src_ws = src_wb.active
    id_to_row = {}
    for r in range(2, src_ws.max_row + 1):
        iid = src_ws.cell(row=r, column=1).value
        if iid:
            id_to_row[str(iid).strip()] = r

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WA_Annotation_FINAL"

    headers = ["item_id", "role_id / target_role", "item_text", "gold_label",
               "weak_subtype", "score (0-3)", "overlap_type",
               "required_capability_blocks", "annotation_rationale", "reviewer_decision"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = BORDER
    ws.freeze_panes = "A2"
    widths = [10, 22, 55, 14, 10, 9, 20, 26, 45, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    row_out = 2
    for i in range(101, 151):
        qid = f"Q_{i}"
        src_row = id_to_row[qid]
        role = src_ws.cell(row=src_row, column=2).value
        text = src_ws.cell(row=src_row, column=3).value
        label = src_ws.cell(row=src_row, column=4).value
        subtype, score, overlap_type, blocks, rationale, decision = ANNOTATIONS[qid]

        values = [qid, role, text, label, subtype, score, overlap_type, blocks, rationale, decision]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row_out, column=c, value=v)
            cell.font = copy(BODY_FONT)
            cell.alignment = copy(WRAP)
            cell.border = copy(BORDER)

        fill = REWRITE_FILL if "rewritten" in decision else (WAD_FILL if subtype == "WA-D" else WAF_FILL)
        for c in range(1, len(headers) + 1):
            ws.cell(row=row_out, column=c).fill = copy(fill)
        row_out += 1

    summary_row = row_out + 1
    waf_n = sum(1 for v in ANNOTATIONS.values() if v[0] == "WA-F")
    wad_n = sum(1 for v in ANNOTATIONS.values() if v[0] == "WA-D")
    rewritten_n = sum(1 for v in ANNOTATIONS.values() if "rewritten" in v[5])
    ws.cell(row=summary_row, column=1, value="SUMMARY (FINAL)").font = Font(name="Arial", size=10, bold=True)
    lines = [
        f"WA-F (Transferable Foundation Overlap): {waf_n} items",
        f"WA-D (Domain-Adjacent Engineering): {wad_n} items",
        f"Total: {waf_n + wad_n} (target 25/25)",
        f"Newly rewritten items (amber): {rewritten_n} -- replaced rejects, length-matched to WA-D band (~30-45 words)",
        "9 rejected originals moved to Adjacent_Role_Challenge_Set_v2.xlsx",
        "Legend: green=WA-F, blue=WA-D, amber=newly rewritten replacement.",
    ]

    for i, line in enumerate(lines, start=1):
        ws.cell(row=summary_row + i, column=1, value=line).font = Font(name="Arial", size=10, italic=True)

    wb.save(OUT_PATH)
    print(f"OK, Saved {OUT_PATH}  WA-F={waf_n} WA-D={wad_n} total={waf_n+wad_n} rewritten={rewritten_n}")


if __name__ == "__main__":
    main()
