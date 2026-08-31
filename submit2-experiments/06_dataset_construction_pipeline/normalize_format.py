"""
    Quick formatting fix: Normalize the item_text in the complete dataset containing 200 entries
"""
import re
import openpyxl
from copy import copy

MAIN_SRC = "/sessions/practical-upbeat-wright/mnt/outputs/Dataset_items_v4_final.xlsx"
MAIN_OUT = "/sessions/practical-upbeat-wright/mnt/outputs/Dataset_items_v5_normalized.xlsx"

CHALLENGE_SRC = "/sessions/practical-upbeat-wright/mnt/outputs/Adjacent_Role_Challenge_Set_v2.xlsx"
CHALLENGE_OUT = "/sessions/practical-upbeat-wright/mnt/outputs/Adjacent_Role_Challenge_Set_v3.xlsx"

ANNOT_SRC = "/sessions/practical-upbeat-wright/mnt/outputs/Weakly_Aligned_Annotation_FINAL.xlsx"
ANNOT_OUT = "/sessions/practical-upbeat-wright/mnt/outputs/Weakly_Aligned_Annotation_FINAL_v2.xlsx"


def normalize_text(text):
    if text is None:
        return text
    t = str(text)
    t = t.replace("Question:", "")
    t = t.replace("•", "")   # bullet character "•"
    t = t.replace("\n", " ")
    t = re.sub(r"\s+", " ", t).strip()
    return t


def normalize_workbook(src_path, out_path, text_col):
    wb = openpyxl.load_workbook(src_path)
    ws = wb.active
    n_changed = 0
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=text_col)
        if cell.value is None or not isinstance(cell.value, str):
            continue
        # Skip rows where column A does not contain item_id
        if ws.cell(row=r, column=1).value is None:
            continue

        new_val = normalize_text(cell.value)
        if new_val != cell.value:
            n_changed += 1
        cell.value = new_val
    wb.save(out_path)
    print(f"OK, {out_path}: normalized {n_changed} rows")
    return out_path


def main():
    normalize_workbook(MAIN_SRC, MAIN_OUT, text_col=3)
    normalize_workbook(CHALLENGE_SRC, CHALLENGE_OUT, text_col=3)
    normalize_workbook(ANNOT_SRC, ANNOT_OUT, text_col=3)

    # Perform structural feature checks on the main dataset
    print("\n" + "*" * 70)
    print("  Verification: structural features after normalization")
    print("*" * 70)
    wb = openpyxl.load_workbook(MAIN_OUT)
    ws = wb.active
    bad = []
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(row=r, column=1).value
        if iid is None:
            continue
        text = ws.cell(row=r, column=3).value or ""
        has_bullet = "•" in text
        has_question_label = "Question:" in text
        has_newline = "\n" in text
        if has_bullet or has_question_label or has_newline:
            bad.append((iid, has_bullet, has_question_label, has_newline))
    if bad:
        print(f"WARN, {len(bad)} rows still contain a flagged structural marker:")
        for row in bad[:20]:
            print("  ", row)
    else:
        print("OK, No row contains a bullet symbol, 'Question:' label, or newline. "
              "Every item is now a single paragraph.")

    # Random sampling data
    wb_src = openpyxl.load_workbook(MAIN_SRC)
    ws_src = wb_src.active
    print("\nSpot-check (before -> after):")
    for iid_target in ["Q_001", "Q_104", "Q_151"]:
        for r in range(2, ws_src.max_row + 1):
            if ws_src.cell(row=r, column=1).value == iid_target:
                before = ws_src.cell(row=r, column=3).value
                after = ws.cell(row=r, column=3).value
                print(f"\n  {iid_target} BEFORE: {before[:150]!r}...")
                print(f"  {iid_target} AFTER:  {after[:150]!r}...")
                break


if __name__ == "__main__":
    main()
