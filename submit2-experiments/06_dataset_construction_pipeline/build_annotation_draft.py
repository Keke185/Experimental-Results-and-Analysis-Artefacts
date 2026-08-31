
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

SRC_PATH = "/sessions/practical-upbeat-wright/mnt/uploads/Dataset_items_Weakly_Aligned_revised.xlsx"
OUT_PATH = "/sessions/practical-upbeat-wright/mnt/outputs/Weakly_Aligned_Annotation_Draft.xlsx"

ANNOTATIONS = {
    "Q_101": ("WA-F", 1, "Supporting (single)", "Troubleshooting",
              "Troubleshooting methodology for production issues with no obvious single point of failure; a genuine troubleshooting support capability, not scenario dressing.",
              "Keep - WA-F"),
    "Q_102": ("REJECT", 0, "None", "-",
              "A plain user-CRUD interface design exercise, an entry-level task for any backend role; does not map to any supporting-capability category and is only tangentially related because \"it is also software development\".",
              "Move to challenge set"),
    "Q_103": ("WA-F", 1, "Supporting (single)", "Troubleshooting / performance optimisation",
              "Slow-query optimisation is a genuine performance-troubleshooting support capability; generic but technically specific and assessable.",
              "Keep - WA-F"),
    "Q_104": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (distributed state)",
              "Presence tracking for a meeting/communication platform; a scenario-embedded distributed state-consistency problem that does not involve signalling-protocol detail.",
              "Keep - WA-D"),
    "Q_105": ("WA-F", 1, "Supporting (single)", "Distributed service design (load balancing)",
              "Load-balancing principles and health checks; a supporting capability within distributed service design, not scenario dressing.",
              "Keep - WA-F"),
    "Q_106": ("WA-F", 1, "Supporting (single)", "Cross-team technical coordination",
              "Delivering a technical feature in collaboration with PM/QA/infra/frontend; maps clearly to the recognised \"cross-team technical coordination\" category.",
              "Keep - WA-F"),
    "Q_107": ("WA-F", 1, "Supporting (single)", "Troubleshooting / performance optimisation (caching)",
              "Caching strategy and invalidation risk; a performance/reliability support capability, not scenario dressing.",
              "Keep - WA-F"),
    "Q_108": ("WA-F", 2, "Supporting (multiple)", "Troubleshooting + observability (incident management)",
              "A complete incident-response workflow (triage/escalation/rollback/post-mortem), covering both the troubleshooting and observability support capabilities.",
              "Keep - WA-F"),
    "Q_109": ("REJECT", 0, "None", "-",
              "Concerns for a PR code review; a purely engineering-culture/soft-skill question, not mapped to any specific supporting-capability category.",
              "Move to challenge set"),
    "Q_110": ("WA-F", 2, "Supporting (multiple)", "Observability + troubleshooting (performance baseline)",
              "Establishing a performance baseline, load testing, bottleneck identification and reporting; covers both the observability and troubleshooting support capabilities.",
              "Keep - WA-F"),
    "Q_111": ("REJECT", 0, "None", "-",
              "Stakeholder alignment under unclear requirements; a purely soft-skill/product-communication question, not mapped to any technical supporting capability.",
              "Move to challenge set"),
    "Q_112": ("REJECT", 0, "None", "-",
              "Communicating a performance risk to non-technical stakeholders; a purely communication soft skill, not mapped to any technical supporting-capability category.",
              "Move to challenge set"),
    "Q_113": ("WA-D", 2, "Adjacent domain (single)", "API integration / API design",
              "Meeting-scheduling API design (idempotency, time zones, version evolution); scenario-embedded but requires only generic API-design capability, no core protocol knowledge.",
              "Keep - WA-D"),
    "Q_114": ("WA-D", 2, "Adjacent domain (single)", "Async task processing (webhook delivery)",
              "Reliable delivery of meeting-lifecycle webhooks; scenario-embedded but fundamentally a generic asynchronous-event-delivery capability.",
              "Keep - WA-D"),
    "Q_115": ("WA-D", 2, "Adjacent domain (single)", "Cloud deployment / security & authentication (storage + access)",
              "Recording-file metadata service (storage lifecycle, access control); scenario-embedded but generic storage-and-permissions design.",
              "Keep - WA-D"),
    "Q_116": ("WA-D", 2, "Adjacent domain (single)", "Async task processing (async workflow)",
              "Post-meeting asynchronous processing pipeline (transcription/thumbnails/compliance scanning); scenario-embedded but a generic asynchronous task-orchestration capability.",
              "Keep - WA-D"),
    "Q_117": ("REJECT", 0, "None", "-",
              "How to quickly get up to speed in an unfamiliar codebase; a purely soft-skill/professional-practice question, not mapped to any technical supporting capability.",
              "Move to challenge set"),
    "Q_118": ("WA-F", 1, "Supporting (single)", "Troubleshooting",
              "Attributing responsibility for and diagnosing a post-change production issue; the core is still troubleshooting methodology, with communication only a secondary component.",
              "Keep - WA-F"),
    "Q_119": ("WA-D", 2, "Adjacent domain (single)", "Async task processing / API integration (notifications)",
              "Multi-channel delivery of meeting invites/reminders; scenario-embedded but a generic notification-system-design capability.",
              "Keep - WA-D"),
    "Q_120": ("WA-D", 2, "Adjacent domain (single)", "API integration (third-party integration)",
              "Third-party calendar-service integration (OAuth/rate limiting/webhook sync); scenario-embedded but a generic third-party API-integration capability.",
              "Keep - WA-D"),
    "Q_121": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (cross-region)",
              "Cross-region session-state availability and consistency; scenario-embedded but a generic distributed-systems-design capability.",
              "Keep - WA-D"),
    "Q_122": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (concurrency control)",
              "Atomic enforcement of a maximum concurrent-participant count; scenario-embedded but a generic distributed concurrency-control capability.",
              "Keep - WA-D"),
    "Q_123": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (state management)",
              "Waiting-room state-machine design; scenario-embedded but a generic state-management/queueing-design capability.",
              "Keep - WA-D"),
    "Q_124": ("WA-D", 2, "Adjacent domain (single)", "Security & authentication (token service)",
              "Issuance and revocation of short-lived meeting tokens; scenario-embedded but a generic identity token-service-design capability.",
              "Keep - WA-D"),
    "Q_125": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design / API integration (rate limiting)",
              "Distributed rate limiting for meeting create/join APIs; scenario-embedded but a generic rate-limiting-architecture capability.",
              "Keep - WA-D"),
    "Q_126": ("WA-D", 1, "Adjacent domain (single, weak)", "Cloud deployment (feature rollout, weak)",
              "Gradual rollout / feature flags is itself a generic SRE release practice; its link to the meeting scenario is only a surface wrapper and the in-domain anchor is weak -- recommend review on whether to keep.",
              "Review - borderline"),
    "Q_127": ("WA-D", 2, "Adjacent domain (single)", "Security & authentication / observability (audit trail)",
              "Audit logging for meeting-management operations; scenario-embedded but a generic audit-and-compliance-logging capability.",
              "Keep - WA-D"),
    "Q_128": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (reconnect/recovery)",
              "State recovery after disconnection/reconnection; scenario-embedded but a generic distributed state-coordination capability.",
              "Keep - WA-D"),
    "Q_129": ("REJECT", 0, "None", "-",
              "Generic Java exception handling (catch Exception e); falls under the explicitly excluded \"pure Java language feature\" category.",
              "Move to challenge set"),
    "Q_130": ("WA-D", 2, "Adjacent domain (single)", "Observability (health monitoring)",
              "Health monitoring across the session/notification/recording-metadata services; scenario-embedded but a generic observability capability.",
              "Keep - WA-D"),
    "Q_131": ("WA-F", 1, "Supporting (single)", "Observability (audit log schema)",
              "Audit-log database schema design; a genuine observability/compliance support capability, not scenario dressing.",
              "Keep - WA-F"),
    "Q_132": ("WA-F", 1, "Supporting (single)", "Distributed service design (database selection)",
              "Relational vs. non-relational database selection; a genuine data-architecture support capability, not scenario dressing.",
              "Keep - WA-F"),
    "Q_133": ("WA-F", 1, "Supporting (single)", "Cloud deployment / distributed service design (schema migration)",
              "Safe production schema migration; a genuine deployment-safety support capability, not scenario dressing.",
              "Keep - WA-F"),
    "Q_134": ("WA-F", 1, "Supporting (single)", "Distributed service design (read replicas)",
              "Using read replicas to improve performance; a genuine data-replication support capability, not scenario dressing.",
              "Keep - WA-F"),
    "Q_135": ("WA-F", 1, "Supporting (single)", "Troubleshooting (connection pool)",
              "Diagnosing database connection-pool exhaustion; a genuine troubleshooting support capability, not scenario dressing.",
              "Keep - WA-F"),
    "Q_136": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design / security & authentication (multi-tenancy)",
              "Multi-tenant isolation and quotas for a meeting backend; scenario-embedded but a generic multi-tenant-architecture capability.",
              "Keep - WA-D"),
    "Q_137": ("WA-F", 1, "Supporting (single)", "Troubleshooting (performance regression)",
              "Systematic investigation of response-time regression; a genuine troubleshooting support capability, not scenario dressing.",
              "Keep - WA-F"),
    "Q_138": ("WA-D", 2, "Adjacent domain (single)", "Async task processing / observability (event stream)",
              "Event-stream design for participant state changes; scenario-embedded but a generic event-driven-architecture capability.",
              "Keep - WA-D"),
    "Q_139": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (coordination state)",
              "Coordinating screen-share control (explicitly excludes media-transport detail); scenario-embedded but a generic distributed-coordination capability.",
              "Keep - WA-D"),
    "Q_140": ("WA-F", 1, "Supporting (single, weak)", "Security & authentication (user management, weak)",
              "Enterprise user-management module including role assignment and audit logging; touches security & authentication, but is still largely an entry-level CRUD exercise overall -- recommend review.",
              "Review - borderline"),
    "Q_141": ("REJECT", 0, "None", "-",
              "Mediating a technical disagreement between two engineers; a purely soft-skill/judgement question, not mapped to a specific supporting-capability category.",
              "Move to challenge set"),
    "Q_142": ("WA-F", 1, "Supporting (single)", "Observability (ops handover docs)",
              "Operations handover documentation (monitoring/alerting/rollback procedure); a genuine observability/operations support capability.",
              "Keep - WA-F"),
    "Q_143": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (overload protection)",
              "Gateway overload protection ahead of a large meeting; scenario-embedded but a generic rate-limiting/degradation-architecture capability.",
              "Keep - WA-D"),
    "Q_144": ("WA-F", 1, "Supporting (single, weak)", "Troubleshooting / cross-team coordination (incident comms, weak)",
              "Cross-team communication during an incident; consistent in spirit with Q_108/Q_106 but with a heavier communication component -- recommend review on whether it duplicates Q_108.",
              "Review - borderline"),
    "Q_145": ("REJECT", 0, "None", "-",
              "Judging whether a given optimisation is worth the investment; a purely cost-benefit-judgement/soft-skill question, not mapped to a specific supporting-capability category.",
              "Move to challenge set"),
    "Q_146": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (session management)",
              "Backend design for meeting-session create/join/leave; scenario-embedded but a generic session-state-management capability, no signalling-protocol detail.",
              "Keep - WA-D"),
    "Q_147": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design (data modelling)",
              "Metadata model design for real-time audio/video sessions; scenario-embedded but a generic data-modelling capability.",
              "Keep - WA-D"),
    "Q_148": ("WA-D", 2, "Adjacent domain (single) -- needs review", "API integration / distributed service design (protocol gateway)",
              "Dual-protocol session-gateway design; a fairly strongly scenario-embedded item -- needs confirmation on whether it stays at a generic gateway-pattern level without touching concrete protocol (SIP/H.323) conversion detail; the item in this set closest to the Aligned boundary.",
              "Review - borderline"),
    "Q_149": ("WA-D", 2, "Adjacent domain (single)", "Observability (RTC monitoring)",
              "Monitoring-metric design for a real-time session backend; scenario-embedded but a generic observability capability.",
              "Keep - WA-D"),
    "Q_150": ("WA-D", 2, "Adjacent domain (single)", "Distributed service design / cloud deployment (scaling)",
              "Scaling design for a communication platform from hundreds to tens of thousands of concurrent users; scenario-embedded but a generic horizontal-scaling capability.",
              "Keep - WA-D"),
}


#Set style and color
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

REJECT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BORDERLINE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
WAD_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
WAF_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


def main():

    src_wb = openpyxl.load_workbook(SRC_PATH)
    src_ws = src_wb.active
    id_to_row = {}
    for r in range(2, src_ws.max_row + 1):
        iid = src_ws.cell(row=r, column=1).value

        if iid:
            id_to_row[str(iid).strip()] = r

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WA_Annotation_Draft"

    headers = [
        "item_id", "role_id / target_role", "item_text", "gold_label",
        "weak_subtype", "score (0-3)", "overlap_type",
        "required_capability_blocks", "annotation_rationale",
        "reviewer_decision (DRAFT)",
    ]

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = BORDER
    ws.freeze_panes = "A2"

    widths = [10, 22, 55, 14, 10, 9, 20, 26, 45, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    row_out = 2
    for i in range(101, 151):

        qid = f"Q_{i}"
        src_row = id_to_row[qid]
        role = src_ws.cell(row=src_row, column=2).value
        text = src_ws.cell(row=src_row, column=3).value
        label = src_ws.cell(row=src_row, column=4).value
        subtype, score, overlap_type, blocks, rationale, decision = ANNOTATIONS[qid]

        values = [qid, role, text, label, subtype, score, overlap_type, blocks, rationale, decision]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row_out, column=c, value=v)
            cell.font = copy(BODY_FONT)
            cell.alignment = copy(WRAP)
            cell.border = copy(BORDER)

        # Highlight
        if subtype == "REJECT":
            fill = REJECT_FILL
        elif "Review" in decision or "borderline" in overlap_type:
            fill = BORDERLINE_FILL
        elif subtype == "WA-D":
            fill = WAD_FILL
        else:
            fill = WAF_FILL
        for c in range(1, len(headers) + 1):
            ws.cell(row=row_out, column=c).fill = copy(fill)

        row_out += 1

    # Summary
    summary_row = row_out + 1
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(name="Arial", size=10, bold=True)
    reject_n = sum(1 for v in ANNOTATIONS.values() if v[0] == "REJECT")
    waf_n = sum(1 for v in ANNOTATIONS.values() if v[0] == "WA-F")
    wad_n = sum(1 for v in ANNOTATIONS.values() if v[0] == "WA-D")
    borderline_n = sum(1 for v in ANNOTATIONS.values() if "Review" in v[5])

    lines = [
        f"WA-F (Transferable Foundation Overlap): {waf_n} items",
        f"WA-D (Domain-Adjacent Engineering): {wad_n} items",
        f"REJECT (move to challenge set): {reject_n} items",
        f"Flagged borderline / pending review: {borderline_n} items (Q_126, Q_140, Q_144, Q_148)",
        f"Total: {waf_n + wad_n + reject_n} (should be 50)",
        "",
        "Legend: green=WA-F kept, blue=WA-D kept, amber=borderline (pending decision), red=reject -> move to challenge set.",
        "This is a DRAFT annotation; the 'reviewer_decision' column records the proposed action pending confirmation.",
    ]

    for i, line in enumerate(lines, start=1):
        ws.cell(row=summary_row + i, column=1, value=line).font = Font(name="Arial", size=10, italic=(i >= 6))

    wb.save(OUT_PATH)
    print(f"OK, Saved {OUT_PATH}")
    print(f"WA-F={waf_n}  WA-D={wad_n}  REJECT={reject_n}  borderline={borderline_n}  total={waf_n+wad_n+reject_n}")


if __name__ == "__main__":
    main()
