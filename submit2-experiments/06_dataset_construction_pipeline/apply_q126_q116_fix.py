"""
Two text fixes were performed on the v6 dataset with corrected capitalization:
    1. Q_126: Replaced with a domain-rewritten version, supplementing constraints on meeting-level consistency,
     cross-client versions, and cross-instance configuration propagation; answers in SaaS common function flag scenarios cannot be fully reused.

    2. Q_116: Syntax correction, merging the ending fragment into a standard comma-separated clause: "...dead
    letter handling, workflow extension, and user-visible processing status".
"""
import openpyxl

MAIN_SRC = "/sessions/practical-upbeat-wright/mnt/outputs/Dataset_items_v6_casefixed.xlsx"
MAIN_OUT = "/sessions/practical-upbeat-wright/mnt/outputs/Dataset_items_v7_final.xlsx"

ANNOT_SRC = "/sessions/practical-upbeat-wright/mnt/outputs/Weakly_Aligned_Annotation_FINAL_v3.xlsx"
ANNOT_OUT = "/sessions/practical-upbeat-wright/mnt/outputs/Weakly_Aligned_Annotation_FINAL_v4.xlsx"

FIXES = {
    "Q_126": (
        "A new meeting feature should be released gradually to selected "
        "tenants. All participants in the same meeting must observe "
        "consistent behaviour even when they use different client versions "
        "or connect through different backend instances. Design a "
        "feature-flag approach covering tenant targeting, meeting-level "
        "consistency, configuration propagation, backward compatibility, "
        "audit history, rollback, and safe defaults."
    ),
    "Q_116": (
        "After a meeting ends, its recording may require transcription, "
        "thumbnail generation, and compliance scanning. Design the "
        "asynchronous workflow, covering job state and queues, retries and "
        "idempotency, partial failures, dead-letter handling, worker "
        "scaling, and user-visible processing status."
    ),
}

def apply_fixes(src_path, out_path, text_col):
    wb = openpyxl.load_workbook(src_path)
    ws = wb.active
    id_to_row = {}
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(row=r, column=1).value
        if iid:
            id_to_row[str(iid).strip()] = r
    n = 0
    for iid, new_text in FIXES.items():
        if iid in id_to_row:
            ws.cell(row=id_to_row[iid], column=text_col).value = new_text
            n += 1
        else:
            print(f"WARN, {iid} not found in {src_path}")
    wb.save(out_path)
    print(f"OK, {out_path}: applied {n} fixes")

def main():
    apply_fixes(MAIN_SRC, MAIN_OUT, text_col=3)
    apply_fixes(ANNOT_SRC, ANNOT_OUT, text_col=3)

    #Update the explanation for the comment Q_126 in the comment file.
    wb = openpyxl.load_workbook(ANNOT_OUT)
    ws = wb.active
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Q_126":
            ws.cell(row=r, column=7).value = "Adjacent domain (single)"
            ws.cell(row=r, column=9).value = (
                "Gradual rollout / feature flags; after the rewrite it now includes within-meeting "
                "cross-client-version behavioural consistency, cross-backend-instance configuration "
                "propagation, and meeting-level configuration constraints -- the in-domain anchor has "
                "been strengthened, and the item is no longer interchangeable with a generic SaaS scenario."
            )
            ws.cell(row=r, column=10).value = "Keep - WA-D (rewritten, anchor fixed)"
            break
    wb.save(ANNOT_OUT)
    print(f"OK, Updated Q_126 annotation rationale in {ANNOT_OUT}")

    #verification
    wb2 = openpyxl.load_workbook(MAIN_OUT)
    ws2 = wb2.active
    for r in range(2, ws2.max_row + 1):
        if ws2.cell(row=r, column=1).value in ("Q_116", "Q_126"):
            print(ws2.cell(row=r, column=1).value, "->", ws2.cell(row=r, column=3).value)

if __name__ == "__main__":
    main()
