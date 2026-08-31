"""
The final assembly is completed using a five-step framework:

1. Baseline file: `Dataset_items_Weakly_Aligned_revised.xlsx`

2. 9 obsolete general entries (Q_102, 109, 111, 112, 117, 129, 140, 141, 145)

3. Export the 9 original texts as independent "Adjacent Technical Role Challenge Sets"

4. Highlight the 9 modified cells in the main file in yellow.

Output: 50 weakly aligned entries (25 WA-D + 25 WA-F), the main dataset maintains 200 rows;
 9 additional challenge sets are generated. `target_role` and `gold_label` remain unchanged.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy

SRC_PATH = "/sessions/practical-upbeat-wright/mnt/uploads/Dataset_items_Weakly_Aligned_revised.xlsx"
MAIN_OUT_PATH = "/sessions/practical-upbeat-wright/mnt/outputs/Dataset_items_v4_final.xlsx"
CHALLENGE_OUT_PATH = "/sessions/practical-upbeat-wright/mnt/outputs/Adjacent_Role_Challenge_Set_v2.xlsx"

NEW_WA_F = {
    # API integration
    "Q_102": (
        "Your backend integrates with an external identity-verification API "
        "run by a vendor.\n"
        "Question: Design this integration.\n"
        "•Authentication with the vendor's API\n"
        "•Handling vendor latency or timeouts\n"
        "•Avoiding duplicate verification requests\n"
        "•Testing safely before production",
        "API集成 API integration",
    ),
    # API integration, second angle
    "Q_109": (
        "One backend capability must be exposed as an API for both an "
        "internal admin tool and an external partner.\n"
        "Question: Design this API to serve both well.\n"
        "•Versioning for independent evolution\n"
        "•Communicating breaking changes\n"
        "•Handling malformed input\n"
        "•API layer vs. client responsibilities",
        "API集成 API integration",
    ),
    # Security & authentication
    "Q_111": (
        "Your backend must authenticate both human users via password and "
        "machine clients via API keys.\n"
        "Question: Design this authentication model.\n"
        "•Routing the two auth flows\n"
        "•Storing and rotating API keys\n"
        "•Scoping each caller's permissions\n"
        "•Revoking access quickly if compromised",
        "安全与认证 Security & authentication",
    ),
    # Security & authentication, second angle
    "Q_112": (
        "Your team discovers a backend service has been logging sensitive "
        "user data in plaintext for months.\n"
        "Question: Describe your response.\n"
        "•Stopping the exposure immediately\n"
        "•Assessing impact of logged data\n"
        "•Preventing this mistake going forward\n"
        "•Deciding what to disclose, and to whom",
        "安全与认证 Security & authentication",
    ),
    # Async task processing
    "Q_117": (
        "A backend feature must generate a large report that takes several "
        "minutes.\n"
        "Question: Design this asynchronously.\n"
        "•Handing off work without blocking the request\n"
        "•Notifying the user when ready\n"
        "•Handling a job that fails partway\n"
        "•Preventing duplicate concurrent generation",
        "异步任务处理 Async task processing",
    ),
    # Async task processing, second angle
    "Q_129": (
        "Your backend runs scheduled background jobs (cleanup, reporting, "
        "syncing) via a job scheduler.\n"
        "Question: Design for reliability.\n"
        "•Preventing concurrent duplicate runs\n"
        "•Handling jobs that repeatedly fail\n"
        "•Monitoring whether jobs keep pace\n"
        "•Changing a schedule without missing or duplicating a run",
        "异步任务处理 Async task processing",
    ),
    # Cloud deployment
    "Q_140": (
        "Your team is migrating a backend service from an on-premises data "
        "center to the cloud.\n"
        "Question: Plan this migration.\n"
        "•Sequencing what moves first vs. last\n"
        "•Keeping the service available throughout\n"
        "•Validating cloud behavior matches on-prem\n"
        "•Rolling back if migration fails badly",
        "云端部署 Cloud deployment",
    ),

    # Distributed service design
    "Q_141": (
        "Two backend services must stay in sync: one updates a record, the "
        "other must eventually reflect it.\n"
        "Question: Design this synchronization.\n"
        "•Sync vs. eventual consistency\n"
        "•Resolving conflicting updates\n"
        "•Recovering from downtime during an update\n"
        "•Verifying the two haven't drifted apart",
        "Distributed service design",
    ),

    # Cross-team technical coordination
    "Q_145": (
        "Your team depends on an API owned by another team, which is "
        "planning a breaking change.\n"
        "Question: Manage this dependency.\n"
        "•Learning about the change early enough\n"
        "•Negotiating a workable migration timeline\n"
        "•Testing against the change before it ships\n"
        "•Handling a slipping timeline",
        "Cross-team technical coordination",
    ),
}

# Set display color and style
HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def main():
    wb = openpyxl.load_workbook(SRC_PATH)
    ws = wb.active
    print(f"OK, Loaded '{ws.title}', dims={ws.dimensions}")

    id_to_row = {}
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(row=r, column=1).value
        if iid:
            id_to_row[str(iid).strip()] = r

    # Get the raw text of 9 rejected projects
    challenge_rows = []
    for item_id in NEW_WA_F:
        r = id_to_row[item_id]
        orig_text = ws.cell(row=r, column=3).value
        role = ws.cell(row=r, column=2).value
        label = ws.cell(row=r, column=4).value
        meta = ws.cell(row=r, column=5).value
        challenge_rows.append((item_id, role, orig_text, label, meta))

    # Overwrite item_text with the new WA-F content.
    changed = []
    for item_id, (new_text, category) in NEW_WA_F.items():
        r = id_to_row[item_id]
        label = ws.cell(row=r, column=4).value
        if label != "Weakly Aligned":
            print(f"WARN, {item_id} row {r} label={label!r}, expected Weakly Aligned , skipping")
            continue
        cell = ws.cell(row=r, column=3)
        cell.value = new_text
        cell.fill = copy(HIGHLIGHT_FILL)
        changed.append((item_id, r, category))

    print(f"OK, Replaced + highlighted {len(changed)} cells (new WA-F items)")

    wb.save(MAIN_OUT_PATH)
    print(f"OK, Saved main dataset: {MAIN_OUT_PATH}")

    # Create a challenge set workbook
    cwb = openpyxl.Workbook()
    cws = cwb.active
    cws.title = "Adjacent_Role_Challenge_Set"
    headers = ["item_id", "role_id / target_role", "item_text (original, unmodified)",
               "gold_label (as originally assigned)", "metadata / trap_type",
               "removal_reason"]
    for c, h in enumerate(headers, start=1):
        cell = cws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = BORDER
    cws.freeze_panes = "A2"
    widths = [10, 22, 60, 18, 20, 45]
    for i, w in enumerate(widths, start=1):
        cws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    reasons = {
        "Q_102": "A plain user-CRUD interface, not mapped to any supporting-capability category and only tangentially related because \"it is also software development\".",
        "Q_109": "Concerns for a PR code review; a purely engineering-culture/soft-skill question.",
        "Q_111": "Stakeholder alignment under unclear requirements; a purely soft-skill/product-communication question.",
        "Q_112": "Communicating a performance risk to non-technical stakeholders; a purely communication soft skill.",
        "Q_117": "How to quickly get up to speed in an unfamiliar codebase; a purely soft-skill/professional-practice question.",
        "Q_129": "Generic Java exception handling (catch Exception e); falls under the explicitly excluded \"pure Java language feature\" category.",
        "Q_140": "Enterprise user-management module; still fundamentally an entry-level CRUD exercise, not meeting the supporting-capability-mapping requirement.",
        "Q_141": "Mediating a technical disagreement between two engineers; a purely soft-skill/judgement question.",
        "Q_145": "Judging whether a given optimisation is worth the investment; a purely cost-benefit-judgement/soft-skill question.",
    }

    row_out = 2
    for item_id, role, orig_text, label, meta in challenge_rows:
        values = [item_id, role, orig_text, label, meta, reasons.get(item_id, "")]
        for c, v in enumerate(values, start=1):
            cell = cws.cell(row=row_out, column=c, value=v)
            cell.font = copy(BODY_FONT)
            cell.alignment = copy(WRAP)
            cell.border = copy(BORDER)
        row_out += 1

    note_row = row_out + 1
    cws.cell(row=note_row, column=1,
             value="Adjacent Technical Role Challenge Set -- measures whether the system mistakes "
                   "generic backend competence for partial target-role alignment just because the item "
                   "came from an adjacent technical role. Not part of the primary 3-class experiment.").font = \
        Font(name="Arial", size=10, italic=True)

    cwb.save(CHALLENGE_OUT_PATH)
    print(f"OK, Saved challenge set: {CHALLENGE_OUT_PATH} ({len(challenge_rows)} items)")

    # Verification
    wb2 = openpyxl.load_workbook(MAIN_OUT_PATH)
    ws2 = wb2.active
    ok = 0
    for item_id, r, category in changed:
        cell = ws2.cell(row=r, column=3)
        expected = NEW_WA_F[item_id][0]
        fill_ok = cell.fill.start_color.rgb in ("00FFFF00", "FFFFFF00")
        text_ok = cell.value == expected
        if fill_ok and text_ok:
            ok += 1
        else:
            print(f"VERIFY WARN, {item_id}: text_ok={text_ok} fill_ok={fill_ok}")
    print(f"VERIFY, {ok}/{len(changed)} new WA-F cells confirmed correct.")

    # Confirm categories now covered
    print("\nNew WA-F category coverage added:")
    for item_id, (text, cat) in NEW_WA_F.items():
        print(f"  {item_id}: {cat}")

    print(f"\n Total rows in main output: {ws2.max_row - 1}")


if __name__ == "__main__":
    main()
