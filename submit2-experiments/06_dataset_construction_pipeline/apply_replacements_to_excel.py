"""
Apply the finalized 33-item Weakly Aligned redesign directly to the user's
uploaded 'Dataset items.xlsx', highlighting every changed item_text cell.

Only column C (item_text) is modified, and only for the 33 item_ids in
REPLACEMENTS below. Column B (role_id/target_role), column D (gold_label),
and column E (metadata/trap_type) are left completely untouched for every
row in the entire 200-row dataset, matching the file's original conventions
except for a yellow highlight fill applied to changed cells so
the advisor/user can see exactly what changed at a glance.
"""
import openpyxl
from openpyxl.styles import PatternFill
from copy import copy

SRC_PATH = "/sessions/practical-upbeat-wright/mnt/uploads/Dataset items.xlsx"
OUT_PATH = "/sessions/practical-upbeat-wright/mnt/outputs/Dataset_items_v2_highlighted.xlsx"

REPLACEMENTS = {
    "Q_104":
        "A backend service maintains persistent WebSocket connections with clients "
        "to support real-time signaling for an audio/video application.\n"
        "Question: Explain how you would design:\n"
        "•Connection establishment and authentication\n"
        "•Heartbeat/keep-alive handling\n"
        "•Reconnection after a dropped connection\n"
        "•Cleanup when a client disconnects unexpectedly\n"
        "What state would need to be reconciled after a client reconnects mid-session?",

    "Q_105":
        "An application needs to show which users are currently online and available "
        "to join a call.\n"
        "Question: Design a presence-tracking service.\n"
        "Explain:\n"
        "•How presence state is updated and propagated\n"
        "•How you would handle stale presence data after an ungraceful disconnect\n"
        "•Scalability considerations as the number of concurrent users grows\n"
        "•How presence data would be consumed by client applications",

    "Q_109":
        "A communications platform needs to store and later retrieve recordings of "
        "past sessions.\n"
        "Question: Design the backend for storing, indexing, and retrieving session "
        "recordings.\n"
        "Explain:\n"
        "•Storage format and lifecycle (hot vs. archival storage)\n"
        "•Metadata needed to search and retrieve recordings later\n"
        "•Access control for who can retrieve a given recording\n"
        "•How you would handle retention and deletion policies",

    "Q_111":
        "Users need to be notified of an incoming call or meeting invite across web "
        "and mobile clients.\n"
        "Question: Design a notification delivery service for real-time call/meeting "
        "invites.\n"
        "Explain:\n"
        "•How you would route a notification to the correct set of client devices\n"
        "•Handling delivery failures and retries\n"
        "•Avoiding duplicate or delayed notifications\n"
        "•How this integrates with the session-creation backend",

    "Q_112":
        "A public API is used by third-party applications to create and manage "
        "real-time sessions.\n"
        "Question: Design a rate-limiting strategy for this signaling API.\n"
        "Explain:\n"
        "•How you would define and enforce rate limits per client/API key\n"
        "•How you would handle legitimate traffic bursts (e.g. many users joining "
        "a large meeting at once)\n"
        "•Trade-offs between strict limits and service protection\n"
        "•How you would communicate rate-limit status back to API consumers",

    "Q_114":
        "After a session ends, several background tasks need to run: generating a "
        "transcript, computing analytics, and archiving the recording.\n"
        "Question: Design a queueing system for this post-session processing.\n"
        "Explain:\n"
        "•How you would sequence or parallelize these tasks\n"
        "•How you would handle a task that fails partway through\n"
        "•How you would avoid reprocessing the same session twice\n"
        "•How you would monitor the health of the processing pipeline",

    "Q_115":
        "Your backend needs to integrate with a third-party video/audio SDK to add "
        "real-time communication features to your product, without building the "
        "underlying media infrastructure yourselves.\n"
        "Question: Explain how you would design the integration layer between your "
        "backend and the third-party SDK.\n"
        "•What configuration and credentials management is needed\n"
        "•How you would abstract the SDK so your application isn't tightly "
        "coupled to one vendor\n"
        "•How you would handle SDK-reported errors or session failures\n"
        "•What you would test before relying on this integration in production",

    "Q_117":
        "A real-time communication service needs to operate across multiple "
        "geographic regions for lower latency, while keeping session state "
        "consistent.\n"
        "Question: Explain how you would replicate session state across regions.\n"
        "•What data needs to be replicated versus kept region-local\n"
        "•How you would handle a user connecting to a different region "
        "mid-session\n"
        "•Consistency trade-offs you would accept\n"
        "•How you would detect and recover from a region outage",

    "Q_119":
        "A conferencing application allows one participant to share their screen "
        "while others view it.\n"
        "Question: Design the backend coordination logic for screen sharing, "
        "separate from the media transport itself.\n"
        "•How you would track which participant currently holds sharing control\n"
        "•How you would handle a request to take over sharing\n"
        "•How you would notify all participants of a sharing state change\n"
        "•How you would handle the sharer disconnecting unexpectedly",

    "Q_122":
        "External systems need to be notified in near-real-time when session events "
        "occur (e.g. call started, participant joined, call ended).\n"
        "Question: Design a webhook delivery system for these events.\n"
        "•How you would guarantee at-least-once delivery\n"
        "•How you would handle a subscriber endpoint that is slow or unavailable\n"
        "•How you would let subscribers verify the authenticity of a webhook\n"
        "•How you would let subscribers filter which events they receive",

    "Q_125":
        "Users can schedule future meetings through an API, and some integrations "
        "create meetings in bulk on behalf of many users.\n"
        "Question: Design the API and its rate-limiting behavior for meeting "
        "scheduling.\n"
        "•How you would prevent bulk scheduling from overwhelming the backend\n"
        "•How you would handle scheduling conflicts or duplicate requests\n"
        "•What validation you would apply to a scheduling request\n"
        "•How you would notify a caller when a scheduling request is throttled",

    "Q_128":
        "A client's network connection drops briefly during an active real-time "
        "session and then recovers.\n"
        "Question: Describe how the backend should handle this scenario.\n"
        "•What session state needs to be preserved during the disconnect window\n"
        "•How the client would resynchronize state upon reconnecting\n"
        "•How you would distinguish a temporary drop from the client actually "
        "leaving\n"
        "•What timeout values you would choose and why",

    "Q_129":
        "A communications platform sells different subscription tiers that allow "
        "different maximum numbers of concurrent session participants.\n"
        "Question: Design the backend logic for enforcing these participant limits.\n"
        "•Where in the request path this check should occur\n"
        "•How you would handle a participant trying to join a session that is "
        "already at capacity\n"
        "•How licensing/seat data would be kept up to date\n"
        "•How you would handle a customer temporarily exceeding their tier "
        "during a grace period",

    "Q_131":
        "An operations team wants a single API to check the real-time health of "
        "active sessions across many backend media-handling instances.\n"
        "Question: Design this monitoring/status API.\n"
        "•What health signals you would aggregate from each instance\n"
        "•How you would handle an instance that stops reporting\n"
        "•How you would summarize fleet-wide health versus per-session detail\n"
        "•How this API would be used to trigger alerts",

    "Q_136":
        "A conferencing platform needs to support configurable conference "
        "control policies -- for example, muting participants on entry, "
        "locking a meeting once it starts, or requiring host approval before "
        "recording begins.\n"
        "Question: Design the backend for enforcing these conference control "
        "policies.\n"
        "•How you would represent and store a policy configuration per "
        "meeting or per organization\n"
        "•How you would enforce a policy consistently regardless of which "
        "client a participant uses\n"
        "•How you would handle a policy being changed while a meeting is "
        "already in progress\n"
        "•How you would audit whether a policy was actually enforced "
        "correctly",

    "Q_137":
        "Clients need a short-lived access token to authenticate when joining a "
        "real-time session.\n"
        "Question: Design the backend service that issues and validates these "
        "tokens.\n"
        "•What claims/information the token should contain\n"
        "•How you would handle token expiry during a long-running session\n"
        "•How you would revoke a token if a session needs to be terminated early\n"
        "•How you would prevent a token from being replayed by an unauthorized "
        "client",

    "Q_140":
        "A regulated enterprise customer requires a detailed audit trail of who "
        "joined, left, and took actions during each session.\n"
        "Question: Design the audit-logging system for this requirement.\n"
        "•What events you would capture and what metadata each entry needs\n"
        "•How you would ensure audit logs cannot be tampered with\n"
        "•How you would support efficient querying of historical audit records\n"
        "•How this interacts with data retention and deletion requirements",

    "Q_141":
        "A conferencing application wants participants to wait in a virtual waiting "
        "room until the host admits them.\n"
        "Question: Design the backend logic for this waiting-room feature.\n"
        "•How you would track who is waiting versus admitted\n"
        "•How you would notify the host of new arrivals\n"
        "•How you would handle a participant leaving the waiting room before "
        "being admitted\n"
        "•How this interacts with the session's overall participant-count "
        "limits",

    "Q_143":
        "A single real-time session depends on several backend subsystems "
        "(signaling, media routing, storage, notifications).\n"
        "Question: Design a health-check endpoint that reports whether a specific "
        "session is fully healthy.\n"
        "•How you would aggregate status from each dependent subsystem\n"
        "•How you would handle a partial failure (some subsystems healthy, "
        "others not)\n"
        "•What response you would return to a client checking session health\n"
        "•How you would avoid this health check becoming a performance "
        "bottleneck itself",

    "Q_145":
        "An API gateway sits in front of your real-time signaling backend and needs "
        "to absorb sudden bursts of traffic, such as many participants joining a "
        "large scheduled event at the same time.\n"
        "Question: Design the gateway's rate-limiting and traffic-shaping policy for "
        "this scenario.\n"
        "•How you would distinguish a legitimate burst from abusive traffic\n"
        "•How you would queue or shed load gracefully rather than failing "
        "outright\n"
        "•How you would communicate back-pressure to upstream clients\n"
        "•How you would validate this policy before a real large-scale event",

    "Q_102":
        "Your backend needs to integrate with a third-party payment provider's "
        "REST API to process transactions.\n"
        "Question: Explain how you would design this integration.\n"
        "•How you would handle authentication with the external API\n"
        "•How you would handle rate limits or downtime on the provider's side\n"
        "•How you would ensure a transaction isn't processed twice if a request "
        "times out\n"
        "•How you would test this integration before going to production",

    "Q_103":
        "A mobile application and a web application both need to consume the same "
        "backend functionality through an API.\n"
        "Question: Explain how you would design this API to serve both clients "
        "well.\n"
        "•How you would version the API to support future changes\n"
        "•How you would document the API for other teams to consume\n"
        "•How you would handle a client sending malformed or unexpected input\n"
        "•How you would decide what belongs in the API layer versus the client",

    "Q_107":
        "Your company relies on several external SaaS APIs (e.g. email delivery, "
        "analytics, storage) that occasionally have outages.\n"
        "Question: Explain how you would design your backend to be resilient to a "
        "failing external dependency.\n"
        "•Timeout and retry strategy\n"
        "•Circuit breaker pattern\n"
        "•Fallback behavior when the dependency is unavailable\n"
        "•How you would monitor the health of these external integrations",

    "Q_113":
        "Your backend API needs to support both first-party client applications "
        "and third-party developers building on your platform.\n"
        "Question: Design the authentication and authorization model.\n"
        "•Difference between authenticating a user versus authenticating an API "
        "client\n"
        "•How you would issue and manage API keys or tokens\n"
        "•How you would scope permissions so a client only accesses what it "
        "should\n"
        "•How you would handle revoking access if a key is compromised",

    "Q_116":
        "A backend service needs to accept user login via username/password as "
        "well as third-party single sign-on (SSO).\n"
        "Question: Explain how you would design this authentication system.\n"
        "•How credentials and sessions would be managed securely\n"
        "•How you would integrate an SSO provider\n"
        "•How you would handle account linking if a user has both login methods\n"
        "•What security risks you would specifically guard against",

    "Q_120":
        "Your team discovers that an internal API has been accidentally exposed "
        "without proper authentication.\n"
        "Question: Describe how you would respond.\n"
        "•Immediate steps to contain the exposure\n"
        "•How you would assess what data or actions may have been affected\n"
        "•How you would prevent this class of mistake from happening again\n"
        "•How you would communicate the incident internally",

    "Q_121":
        "Your application is currently a single monolithic service, and the team "
        "is considering splitting parts of it into separate services.\n"
        "Question: Explain how you would approach this decision.\n"
        "•What criteria you would use to decide which parts to split out\n"
        "•How services would communicate with each other\n"
        "•How you would handle data consistency across service boundaries\n"
        "•What new operational complexity this introduces",

    "Q_123":
        "Multiple backend services need to coordinate to complete a single "
        "business operation that spans all of them.\n"
        "Question: Explain how you would design this coordination.\n"
        "•Synchronous request chaining versus event-driven coordination\n"
        "•How you would handle a partial failure partway through the operation\n"
        "•How you would keep the overall operation observable/traceable\n"
        "•Trade-offs between consistency and availability in this design",

    "Q_124":
        "Your team needs to deploy a backend service to the cloud with the "
        "ability to scale up during peak traffic and scale down afterward.\n"
        "Question: Explain how you would design the deployment.\n"
        "•How you would define auto-scaling triggers\n"
        "•How you would handle stateful vs. stateless components differently\n"
        "•How you would roll out a new version without downtime\n"
        "•How you would control cloud infrastructure costs",

    "Q_126":
        "A backend service currently runs in a single cloud region, and the "
        "business wants better resilience against a regional outage.\n"
        "Question: Explain your approach to improving this.\n"
        "•What it would take to run in multiple regions\n"
        "•How you would handle failover if one region goes down\n"
        "•What data would need to be replicated versus kept region-local\n"
        "•How you would test that failover actually works before you need it",

    "Q_127":
        "Your application needs to send a confirmation email after a user "
        "completes an action, but email delivery is sometimes slow.\n"
        "Question: Explain how you would design this so it doesn't block the "
        "user's request.\n"
        "•How you would decouple the email sending from the main request flow\n"
        "•How you would handle a failed email delivery attempt\n"
        "•How you would avoid sending duplicate emails\n"
        "•How you would monitor whether the email queue is falling behind",

    "Q_130":
        "A backend service intermittently returns 500 errors, but the issue "
        "cannot be reproduced locally.\n"
        "Question: Describe your investigation approach.\n"
        "•What logs, metrics, or traces you would look at first\n"
        "•How you would try to reproduce the issue in a safer environment\n"
        "•How you would narrow down whether the cause is code, infrastructure, "
        "or a dependency\n"
        "•How you would confirm a fix actually resolved the issue",

    "Q_138":
        "Your team is being asked to improve visibility into how a backend "
        "service is performing in production.\n"
        "Question: Explain what you would put in place.\n"
        "•What metrics you would track and why\n"
        "•How you would set up alerting thresholds\n"
        "•How logs and traces would help diagnose issues faster\n"
        "•How you would avoid alert fatigue from too many low-value alerts",
}

HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def main():
    wb = openpyxl.load_workbook(SRC_PATH)
    ws = wb.active
    print(f"OK, Loaded '{ws.title}', dims={ws.dimensions}")

    # Map item_id -> row number by scanning column A (header in row 1).
    id_to_row = {}
    for row_idx in range(2, ws.max_row + 1):
        iid = ws.cell(row=row_idx, column=1).value
        if iid is not None:
            id_to_row[str(iid).strip()] = row_idx

    missing = [iid for iid in REPLACEMENTS if iid not in id_to_row]
    if missing:
        print(f"WARN, item_ids not found, skipped: {missing}")

    changed = []
    for item_id, new_text in REPLACEMENTS.items():
        if item_id not in id_to_row:
            continue
        row_idx = id_to_row[item_id]
        cell = ws.cell(row=row_idx, column=3)  # column C = item_text

        # row/label should be Weakly Aligned before we touch it.
        label = ws.cell(row=row_idx, column=4).value
        if label != "Weakly Aligned":
            print(f"WARN, {item_id} at row {row_idx} has gold_label={label!r}, expected 'Weakly Aligned' -- skipping.")
            continue

        cell.value = new_text
        cell.fill = copy(HIGHLIGHT_FILL)
        changed.append((item_id, row_idx))

    print(f"OK, Updated + highlighted {len(changed)} item_text cells.")

    wb.save(OUT_PATH)
    print(f"OK, Saved to {OUT_PATH}")

    # Reload and confirm after successful verification
    wb2 = openpyxl.load_workbook(OUT_PATH)
    ws2 = wb2.active
    ok = 0
    for item_id, row_idx in changed:
        cell = ws2.cell(row=row_idx, column=3)
        expected = REPLACEMENTS[item_id]
        fill_ok = cell.fill.start_color.rgb in ("00FFFF00", "FFFFFF00")
        text_ok = cell.value == expected
        if fill_ok and text_ok:
            ok += 1
        else:
            print(f"VERIFY WARN, {item_id} row {row_idx}: text_ok={text_ok} fill_ok={fill_ok} fill={cell.fill.start_color.rgb}")
    print(f"VERIFY, {ok}/{len(changed)} changed cells confirmed correct (text + highlight).")

    # Confirm untouched columns for a few sample rows
    sample_ids = ["Q_104", "Q_136", "Q_146", "Q_001", "Q_151"]
    for sid in sample_ids:
        if sid in id_to_row:
            r = id_to_row[sid]
            print(f"  sample {sid} (row {r}): role={ws2.cell(row=r, column=2).value!r}, "
                  f"gold_label={ws2.cell(row=r, column=4).value!r}, "
                  f"metadata={ws2.cell(row=r, column=5).value!r}")

    print(f"\n Total rows in output: {ws2.max_row - 1}")


if __name__ == "__main__":
    main()
