"""
    Fixing the unnatural capitalization of generic nouns in sentences—another shortcut to tag leaks,
    worse than bullet points/issues: a category scan revealed that the capitalized System appeared
    51 times in the "Mismatch" category but none in the Alignment category
"""

import re
import openpyxl

MAIN_SRC = "/sessions/practical-upbeat-wright/mnt/outputs/Dataset_items_v5_normalized.xlsx"
MAIN_OUT = "/sessions/practical-upbeat-wright/mnt/outputs/Dataset_items_v6_casefixed.xlsx"

CHALLENGE_SRC = "/sessions/practical-upbeat-wright/mnt/outputs/Adjacent_Role_Challenge_Set_v3.xlsx"
CHALLENGE_OUT = "/sessions/practical-upbeat-wright/mnt/outputs/Adjacent_Role_Challenge_Set_v4.xlsx"

ANNOT_SRC = "/sessions/practical-upbeat-wright/mnt/outputs/Weakly_Aligned_Annotation_FINAL_v2.xlsx"
ANNOT_OUT = "/sessions/practical-upbeat-wright/mnt/outputs/Weakly_Aligned_Annotation_FINAL_v3.xlsx"

TARGET_WORDS = [
    "System", "Systems", "User", "Users", "Performance", "Management",
    "Optimize", "Architecture", "Solution", "Solutions", "Platform",
]
PATTERN = re.compile(r"\b(" + "|".join(TARGET_WORDS) + r")\b")


def fix_case(text):
    if not text:
        return text

    def repl(m):
        start = m.start()
        word = m.group(0)
        prefix = text[:start]
        is_text_start = (prefix.strip() == "")
        # If this is the first tag in the entire text, it is a sentence beginning tag
        is_after_sentence_end = bool(re.search(r"[.!?]\s+$", prefix))
        if is_text_start or is_after_sentence_end:
            return word  # The legal beginning of a sentence remains unchanged.
        return word.lower()

    return PATTERN.sub(repl, text)


def process(src_path, out_path, text_col):
    wb = openpyxl.load_workbook(src_path)
    ws = wb.active
    n_changed = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is None:
            continue
        cell = ws.cell(row=r, column=text_col)
        if not isinstance(cell.value, str):
            continue
        new_val = fix_case(cell.value)
        if new_val != cell.value:
            n_changed += 1
        cell.value = new_val
    wb.save(out_path)
    print(f"OK, {out_path}: case-fixed {n_changed} rows")


def main():
    process(MAIN_SRC, MAIN_OUT, text_col=3)
    process(CHALLENGE_SRC, CHALLENGE_OUT, text_col=3)
    process(ANNOT_SRC, ANNOT_OUT, text_col=3)

    print("\n" + "*" * 80)
    print("  Validation: The number of times the remaining uppercase letters appear, categorized by type")
    print("*" * 80)
    from collections import Counter
    wb = openpyxl.load_workbook(MAIN_OUT)
    ws = wb.active
    by_word = {}
    for r in range(2, 202):
        label = ws.cell(row=r, column=4).value
        text = ws.cell(row=r, column=3).value or ""
        for w in TARGET_WORDS:
            for m in re.finditer(r"\b" + w + r"\b", text):
                by_word.setdefault(w, Counter())[label] += 1
    if not by_word:
        print("OK, None of the capital letters of the target word appeared.")
    else:
        for w, c in by_word.items():
            print(f"  {w}: {dict(c)}  (remaining , should only be true sentence-initial uses)")

    # Sampling Inspection
    wb_src = openpyxl.load_workbook(MAIN_SRC)
    ws_src = wb_src.active
    print("\n Spot-check (before -> after) on a Mismatched row with heavy System/User/Solution use:")
    for r in range(2, 202):
        if ws_src.cell(row=r, column=1).value == "Q_001":
            print("BEFORE:", ws_src.cell(row=r, column=3).value)
            print("AFTER: ", ws.cell(row=r, column=3).value)


if __name__ == "__main__":
    main()
