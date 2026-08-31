"""
    Style-confound audit .
"""
import re
import statistics as st
import openpyxl

PATH = "/sessions/practical-upbeat-wright/mnt/outputs/Dataset_items_v4_final.xlsx"


WA_D_IDS = {
    "Q_104","Q_113","Q_114","Q_115","Q_116","Q_119","Q_120","Q_121","Q_122",
    "Q_123","Q_124","Q_125","Q_126","Q_127","Q_128","Q_130","Q_136","Q_138",
    "Q_139","Q_143","Q_146","Q_147","Q_148","Q_149","Q_150",
}
WA_F_IDS = {
    "Q_101","Q_102","Q_103","Q_105","Q_106","Q_107","Q_108","Q_109","Q_110",
    "Q_111","Q_112","Q_117","Q_118","Q_129","Q_131","Q_132","Q_133","Q_134",
    "Q_135","Q_137","Q_140","Q_141","Q_142","Q_144","Q_145",
}

SOFT_SKILL_WORDS = [
    "stakeholder", "communicate", "communication", "disagree", "coordinate",
    "coordination", "align", "alignment", "consensus", "convince", "conflict",
    "negotiate", "priorities", "team consensus", "escalate", "escalation",
]
DOMAIN_KEYWORDS = [
    "meeting", "conference", "video", "call", "session", "participant",
    "webrtc", "sip", "h.323", "vmr", "webinar", "real-time communication",
    "signaling", "media", "recording", "waiting room", "tenant",
]

def word_count(text):
    return len(re.findall(r"\S+", text))

def soft_skill_hits(text):
    t = text.lower()
    return sum(1 for w in SOFT_SKILL_WORDS if w in t)

def domain_hits(text):
    t = text.lower()
    return sum(1 for w in DOMAIN_KEYWORDS if w in t)

def summarize(label, values):
    if not values:
        return
    print(f"  {label:28s} n={len(values):3d}  mean={st.mean(values):6.1f}  "
          f"std={st.pstdev(values):5.1f}  min={min(values):5.1f}  max={max(values):6.1f}  "
          f"median={st.median(values):6.1f}")


def main():
    wb = openpyxl.load_workbook(PATH)
    ws = wb.active

    items = []
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(row=r, column=1).value
        label = ws.cell(row=r, column=4).value
        text = ws.cell(row=r, column=3).value
        if iid:
            items.append((iid, label, text))

    print("*" * 78)
    print("  Style-confound audit: word count by class / subtype")
    print("*" * 78)

    groups = {"Aligned": [], "Mismatched": [], "Weakly Aligned (all)": [],
              "  - WA-F": [], "  - WA-D": []}
    soft = {"Aligned": [], "Mismatched": [], "Weakly Aligned (all)": [],
            "  - WA-F": [], "  - WA-D": []}
    dom = {"Aligned": [], "Mismatched": [], "Weakly Aligned (all)": [],
           "  - WA-F": [], "  - WA-D": []}

    for iid, label, text in items:
        wc = word_count(text)
        sk = soft_skill_hits(text)
        dk = domain_hits(text)
        if label == "Aligned":
            groups["Aligned"].append(wc); soft["Aligned"].append(sk); dom["Aligned"].append(dk)
        elif label == "Mismatched":
            groups["Mismatched"].append(wc); soft["Mismatched"].append(sk); dom["Mismatched"].append(dk)
        elif label == "Weakly Aligned":
            groups["Weakly Aligned (all)"].append(wc)
            soft["Weakly Aligned (all)"].append(sk)
            dom["Weakly Aligned (all)"].append(dk)
            if iid in WA_F_IDS:
                groups["  - WA-F"].append(wc); soft["  - WA-F"].append(sk); dom["  - WA-F"].append(dk)
            elif iid in WA_D_IDS:
                groups["  - WA-D"].append(wc); soft["  - WA-D"].append(sk); dom["  - WA-D"].append(dk)

    print("\n Word count:")
    for k, v in groups.items():
        summarize(k, v)

    print("\n Soft-skill keyword hits (0 = purely technical framing):")
    for k, v in soft.items():
        summarize(k, v)

    print("\n Explicit domain-keyword hits (meeting/video/session/etc -- a give-away "
          "for the model even without real capability overlap):")
    for k, v in dom.items():
        summarize(k, v)

    leaked = [iid for iid, label, text in items if iid in WA_F_IDS and domain_hits(text) > 0]
    print(f"\n CHECK, WA-F items containing an explicit domain keyword (should be 0): "
          f"{len(leaked)} -> {leaked}")

    no_leak = [iid for iid, label, text in items if iid in WA_D_IDS and domain_hits(text) == 0]
    print(f"CHECK, WA-D items with NO domain keyword at all (should be 0): "
          f"{len(no_leak)} -> {no_leak}")

    soft_heavy = [iid for iid, label, text in items if iid in WA_F_IDS and soft_skill_hits(text) >= 2]
    print(f"CHECK, WA-F items still soft-skill-heavy (>=2 keyword hits, risk of "
          f"drifting back toward reject criteria): {len(soft_heavy)} -> {soft_heavy}")


if __name__ == "__main__":
    main()
